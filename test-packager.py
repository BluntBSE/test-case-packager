""" for (test in data_frame):
        if(searchAsana(test) == true):
                updateAsana(test)
        if(searchAsana(test) == false:
                addToListOfNewTests(test)
        else:
            addtoListForHumanVerification(test)
            
        ##Afterwards....
        allCNumbersInAsana = getTestByFunctionalArea(area)
        for (CNumber in allCNumbersinAsana):
            if(not in data_frame['ID']):
                addtoListofMissingtestCases(CNumber)
"""


##MBTA Workspace GID: 15492006741476
##Test bed project GID : 1203186680032258

##Remember to add "Automatically Generated" tag GID = '1203191978331220'
##Remember to add "Automatically Updated" tag GID = ''
##Asana records consist of Test: [Area] and optionally, 
##Create a new CSV containing all test cases received for which we received a C number, but there are no C Numbers in our Asana records. These are new test cases
##Create a CSV containing all test cases where we had an Asana record, but did not receive a C number. These are potentially missing test cases.

##Test Script Review Status GID = 1184137262520341
##Reference Review Required Task GID = 1202894410931244
##Reference Approved Task GID = 1202632050244277
##Custom Field 'ID Number' GID = 515737795293097




#Step 1 - Check to see if a task exists in Asana with the defined C number.

import sys
import requests
import os
import pandas
import numpy
import asana
import re ##Regex -- used for matching name string
import json
import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import date



##HELPERS

def AddTasksToWorksheet(tasks, sheet, workbook):
##QA COMPLETE TAG GID = 649069647070258
##HOLD TAG GID = 259956811260129
##STATUS COLUMN HEADER = OPYXL ROW 10, COLUMN 10 
    qa_tag = {'gid': '649069647070258', 'resource_type': 'tag'}
    hold_tag =  {'gid': '259956811260129', 'resource_type': 'tag'}
    rownum = 11
    for row, task in enumerate(tasks):
        ##OPYXL is 1-indexed. This should be the first row of the output spreadsheet you want to fill. TODO: Find the actual header and drop down one instead of writing '11'
      
            if hold_tag not in task['tags']:
                if qa_tag in task['tags']:
                    for ind, field in enumerate(task['custom_fields']):
                        print(field['name'])
                        ##Add review status...
                        if (field['name'] == 'Test Script Review Status'):
                         ##   print(task['custom_fields'][ind]['display_value'])
                            sheet.cell(row = rownum, column = 10).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'ID Number'):
                            sheet.cell(row = rownum, column = 5).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Functional area'):
                            sheet.cell(row = rownum, column = 4).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 3).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']

                    sheet.cell(row = rownum, column = 7).value = task['notes']
                    sheet.cell(row = rownum, column = 6).value = task['name']
                    rownum = rownum + 1
            
            workbook.save('output.xlsm')




   ## sheet.cell(rownum, column=11).value = task



def get_tasks_by_section(section_gid):
    tasklist = []
    tasks = client.tasks.get_tasks({'section':section_gid, 'completed_since':date.today(), 'opt_fields':['name', 'notes', 'tags', 'custom_fields']}, )
    print(tasks)
    for task in tasks:
        tasklist.append(task)
    return tasklist



##Put your token in a 'credentials.py' in the same directory as this script
sys.path.append(os.path.relpath('.\credentials.py'))
from credentials import token

##SIMPLE ASANA AUTHENTICATION##
##Headers to log in as Robert

##Set token equal to a Personal Access Token


client = asana.Client.access_token(token)
workspace = '15492006741476'

##LOAD THE TEMPLATE

template_contents = os.listdir('template')
if(len(template_contents)>1):
    print('!ERROR! More than one file in template folder')
else:
    temp_filename = os.listdir('template')[0]
    temp_filepath = os.path.dirname(os.path.abspath(__file__)) + '/template/'+ temp_filename
   

temp_xlsx=pandas.ExcelFile(temp_filepath)
##Select the 'Document Comments' worksheet
df_temp = pandas.read_excel(temp_xlsx, "Document Comments")



##LOAD THE INPUT SHEET (WE DO THIS TO IMPORT ANY HEADERS SO CONTRACTS CAN CHANGE THEM...)

input_contents = os.listdir('input')
if(len(template_contents)>1):
    print('!ERROR! More than one file in input folder')
else:
    in_filename = os.listdir('input')[0]
    in_filepath = os.path.dirname(os.path.abspath(__file__)) + '/input/'+ in_filename


in_xlsx=pandas.ExcelFile(in_filepath)
##Select the 'Document Comments' worksheet
df_in = pandas.read_excel(in_xlsx, "Document Comments")

##Opyxl is for insertion
temp_opyxl_wb = openpyxl.load_workbook(in_filepath, keep_vba=True)
temp_opyxl_sheet = temp_opyxl_wb['Document Comments'] 
print(temp_opyxl_sheet.cell(row = 10, column = 10 ).value)
 
##Check row 8 of template + input. Are they the same? These are the headers for where we import content. If these don't match exactly, the program will not run.
df_temp_headers = df_temp.iloc[8]
df_in_headers = df_in.iloc[8]


if(df_in_headers.equals(df_temp_headers)):
    print("Matching!")
else:
    print("This program has determined that row 8 of the input and template are different. Does your input format match the template exactly?")
    quit()


print("This program has determined that row 8 of both the input and template match.")


test_project_id = input("Please enter the ID of the test project (E.G: 'Test: System Website' is `1202257161854797` as extracted from https://app.asana.com/0/1202257161854797/list)")

##Seek and register sections named 'General Commetns' and 'Test cases currently under review'. Case specific, etc.

general_gid = None
cases_gid = None

##Validate that there's only one section with the rquired name in the project. Get the GIDs.
test_project_sections = client.sections.find_by_project(test_project_id)
for section in test_project_sections:
    print(section)
    if(section['name'] == 'General comments'):
        if(general_gid != None):
            print("WARNING! Multiple sections found with the name, 'General comments'. Quitting.")
        general_gid = section['gid']
    if(section['name'] == 'Test cases currently under review'):
        if(cases_gid != None):
            print("Warning! Multiple sections labeled 'Test cases under review'. Quitting.")
            quit()
        cases_gid = section['gid']


gen_tasklist = get_tasks_by_section(general_gid)

cases_tasklist = get_tasks_by_section(cases_gid)

total_tasklist = gen_tasklist + cases_tasklist

AddTasksToWorksheet(total_tasklist, temp_opyxl_sheet, temp_opyxl_wb)


